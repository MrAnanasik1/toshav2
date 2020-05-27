from interpreter import TwinInterpreter
import time
import datetime
import random
import logging
import openpyxl
import os
import speech_recognition as sr
import pyttsx3

engine = pyttsx3.init()

logging.basicConfig(format="%(levelname)s %(asctime)s - %(message)s")
today = datetime.datetime.today()

r = sr.Recognizer()
m = sr.Microphone(device_index=1)


class ActionHandler:
    def __init__(self, confidence: float = 0.5):
        self.last_intent = None # последний обработанный интент
        self.last_entities = None # последние обработанные сущности
        self.confidence = confidence # порог принятия решения
    

    @staticmethod
    def get_any_entity(entities: list, name: str, default = None):
        """
        Вернет первую попавшуюся сущность (dict) с указанным именем или default
        :param entities: список сущностей
        :param name: имя сущности, кранится по ключу 'entity'
        :param default:
        :return:
        """
        for item in entities:
            if item.get('entity','') == name:
                return item
        return default

    @staticmethod
    def get_all_entities(entities: list, name: str):
        """
        Вернет все сущности с указанным именем
        :param entities: список сущностей
        :param name: имя сущности, кранится по ключу 'entity'
        :return:
        """
        result = list()
        for item in entities:
            if item.get('entity', '') == name:
                result.append(item)
        return result

    def process(self,intent, entities):
        """
        Основной метод, который обрабатывает входные намерения
        :param intent:
        :param entities:
        :return:
        """
        # вытаскиваем имя намерения и confidence
        name = intent.get("name","default")
        confidence = intent.get("confidence",0)
        if confidence > self.confidence:
            # если бот уверен в намеренье
            # нахомим метод обработчик, если не найдем - будет вызвана method_not_implemented
            handler = getattr(self, "on_{name}".format(name=name).lower(), self.method_not_implemented)
        else:
            # бот не уверен в намеренье
            handler = self.on_default
        # вызываем метод
        result = handler(intent, entities)
        # cохраняем намерение и сущности
        self.last_intent = intent
        self.last_entities = entities
        # вернуть результат
        return result

    def method_not_implemented(self,intent, entities):
        """
        Как обрабатвется намерение, если для него нет метода обработчика
        :param intent:
        :param entities:
        :return:
        """
        name =  intent.get("name", "")
        confidence = intent.get("confidence", 0)
        msg = "Method on_{intent_name} not implemented".format(intent_name = name)
        logging.warning(msg)
        return "Intent: {name}, confidence: {confidence}".format(name = name, confidence = confidence)

    def on_default(self,intent, entities):
        """
        Как обрабатывается default намерение, оно приходит от ТВИНА если он не понял намерение
        :param intent:
        :param entities:
        :return:
        """

        return random.choice(["Не понял Вас, повторите", "Что-что? Повторите еще раз"])

    def on_twin_greeting(self, intent, entities):
        """
        Как бот обрабатывет приветствие twin_greeting
        :param intent:
        :param entities:
        :return:
        """
        phrases1 = ["Добрый день!","Здравствуйте!","Приветствую вас!"]
        phrases2 = ["Какой у вас вопрос?", "Меня зовут Тоша. Чем я могу помочь?",""]
        return "{p1} {p2}".format(p1 = random.choice(phrases1), p2 = random.choice(phrases2))



    def on_thanks(self, intent, entities):
        """
        Как бот обрабатывет благодарность
        :param intent:
        :param entities:
        :return:
        """

        return random.choice(["Пожалуйста, рад помочь", "Всегда пожалуйста, рад помочь", "Пожалуйста"])


    def on_twin_goodbye(self, intent, entities):
        """
        Как бот обрабатывет когда с ним прощаются twin_goodbye
        :param intent:
        :param entities:
        :return:
        """
        return random.choice(["Всего доброго", "До свидания", "Всегда рад помочь"])
    

    def on_twin_repeat(self, intent, entities):
        """
        Как бот обрабатвает просьбу повторить twin_repeat
        :param intent:
        :param entities:
        :return:
        """
        return self.process(self.last_intent,self.last_entities)

    def on_twin_way(self,intent, entities):
        """
        Как обрабатываем просьбу "справшивают путь"
        :param intent:
        :param entities:
        :return:
        """
        ways = list()
        # обрабатываем каждую сущость, которая может ходить с данным интентом
        if self.get_any_entity(entities,'swimming_pool'):
            ways.append("идите прямо"), 
        if self.get_any_entity(entities,'music_room'):
            ways.append("Поднимитесь на второй этаж и идите налево, вторая дверь")
        if self.get_any_entity(entities,'teaching_room'):
            ways.append("Поднимитесь на второй этаж и идите налево до конца коридора, третья дверь после групп")
        if self.get_any_entity(entities,'laundry'):
            ways.append("идите налево по коридору, третья дверь")
        if self.get_any_entity(entities,'medical_office'):
            ways.append("идите налево по коридору, четвертая дверь")
        if self.get_any_entity(entities,'HR'):
            ways.append("Поднимитесь на второй этаж и идите налево до конца коридора, первая дверь после групп")
        if self.get_any_entity(entities,'accounting'):
            ways.append("Поднимитесь на второй этаж и идите налево до конца коридора, четвертая дверь после групп")
        if self.get_any_entity(entities,'food_block'):
            ways.append("идите налево до железной двери и поверните направо перед ней")
        if self.get_any_entity(entities,'manager'):
            ways.append("Поднимитесь на второй этаж и идите налево до конца коридора, вторая дверь после групп ")
        if self.get_any_entity(entities,'gym'):
            ways.append("Поднимитесь на второй этаж и идите налево, первая дверь")
        if self.get_any_entity(entities,'group_1'):
            ways.append("идите направо по коридору, первая дверь")
        if self.get_any_entity(entities,'group_2'):
            ways.append("идите направо по коридору, вторая дверь")
        if self.get_any_entity(entities,'group_3'):
            ways.append("идите направо по коридору, третья дверь")
        if self.get_any_entity(entities,'group_4'):
            ways.append("идите направо по коридору, четвертая дверь")
        if self.get_any_entity(entities,'group_5'):
            ways.append("идите налево по коридору, первая дверь")
        if self.get_any_entity(entities,'group_6'):
            ways.append("идите налево по коридору, вторая дверь")
        if self.get_any_entity(entities,'group_7'):
            ways.append("Поднимитесь на второй этаж и идите направо, первая дверь ")
        if self.get_any_entity(entities,'group_8'):
            ways.append("Поднимитесь на второй этаж и идите направо, вторая дверь ")
        if self.get_any_entity(entities,'group_9'):
            ways.append("Поднимитесь на второй этаж и идите направо, третья дверь ")
        if self.get_any_entity(entities,'group_10'):
            ways.append("Поднимитесь на второй этаж и идите направо, четвертая дверь ")
        if self.get_any_entity(entities,'group_11'):
            ways.append("Поднимитесь на второй этаж и идите налево, первая дверь")
        if self.get_any_entity(entities,'group_12'):
            ways.append("Поднимитесь на второй этаж и идите налево, вторая дверь")
            
       
        #  формируем ответ
        if len(ways) > 0:
            return ". ".join(ways)
        else:
            return self.on_default(intent,entities)

    def on_ask_phone(self,intent, entities):
        """
        Как обрабатываем просьбу "спрашивать телефон"
        :param intent:
        :param entities:
        :return:
        """
        wb = openpyxl.load_workbook('tosha.xlsx')
        sheet = wb['Телефон']
        
        phone = list()
        # обрабатываем каждую сущость, которая может ходить с данным интентом
        if self.get_any_entity(entities,'accounting'):
            phone.append(str(sheet['B2'].value))
        if self.get_any_entity(entities,'manager'):
            phone.append(str(sheet['B3'].value))
        if self.get_any_entity(entities,'manager_chores'):
            phone.append(str(sheet['B4'].value))
        if self.get_any_entity(entities,'manager_education'):
            phone.append(str(sheet['B5'].value))
        if self.get_any_entity(entities,'medical_office'):
            phone.append(str(sheet['B6'].value))
        if self.get_any_entity(entities,'HR'):
            phone.append(str(sheet['B7'].value))
        
        #  формируем ответ
        if len(phone) > 0:
            return ". ".join(phone)
        else:
            return self.on_default(intent,entities)

    def on_ask_menu(self,intent, entities):
        """
        :param intent:
        :param entities:
        :return:
        """
        wb = openpyxl.load_workbook('tosha.xlsx')
        sheet = wb['Меню']
        def prepare_menu(date: str, times: list):
            """
            Сходить в БД, получить меню
            :param date:
            :param times:
            :return:
            """
            now_date = datetime.datetime.now().strftime('%Y-%m-%d')
            if date != now_date:
                return "Меню на {date} нет".format(date=date)
            return "Меню {times}".format(date=date,times = ", ".join(times))
        # вытаскиваем дату или ставим текущую
        now_date = datetime.datetime.now().strftime('%Y-%m-%d')
        time_entity = self.get_any_entity(entities,'time',dict())
        date = time_entity.get('value',now_date).split('T')[0]
        # смотрим какое время хотят
        TIMES = (
            ("time_breakfast",sheet['B3'].value),
            ("time_dinner",sheet['B4'].value),
            ("time_afternoon_tea",sheet['B5'].value),
            ("time_supper",sheet['B6'].value)
        )
        times = list()
        for e,v in TIMES:
            if self.get_any_entity(entities,e):
                times.append(v)
        if len(times) < 1:
            times = [item[1] for item in TIMES]
        # возвращаем результат
        return prepare_menu(date,times)

    
    def on_ask_plan(self,intent, entities):
        """
        :param intent:
        :param entities:
        :return:
        """

        wb = openpyxl.load_workbook('tosha.xlsx')
        sheet = wb['Мероприятие']
        def prepare_plan(date: str, times: list):
            """
            Сходить в БД, получить планы
            :param date:
            :param times:
            :return:
            """
            now_date = datetime.datetime.now().strftime('%Y-%m-%d')
            if date != now_date:
                return "Мероприятия на {date} нет".format(date=date)
            return "Мероприятие {times}".format(date=date,times = ", ".join(times))
        # вытаскиваем дату или ставим текущую
        now_date = datetime.datetime.now().strftime('%Y-%m-%d')
        time_entity = self.get_any_entity(entities,'time',dict())
        date = time_entity.get('value',now_date).split('T')[0]
        # смотрим какую группу хотят
        TIMES = (
            ("group_1",sheet['B2'].value),
            ("group_2",sheet['B3'].value),
            ("group_3",sheet['B4'].value),
            ("group_4",sheet['B5'].value),
            ("group_5",sheet['B6'].value),
            ("group_6",sheet['B7'].value),
            ("group_7",sheet['B8'].value),
            ("group_8",sheet['B9'].value),
            ("group_9",sheet['B10'].value),
            ("group_10",sheet['B11'].value),
            ("group_11",sheet['B12'].value),
            ("group_12",sheet['B13'].value)
        )
        times = list()
        for e,v in TIMES:
            if self.get_any_entity(entities,e):
                times.append(v)
        if len(times) < 1:
            times = [item[1] for item in TIMES]
        # возвращаем результат
        return prepare_plan(date,times)

class DialogTracker:
    """

    """
    def __init__(self,interpreter, action_handler):
        self.interpreter = interpreter
        self.action_handler = action_handler

    def predict(self, text: str) ->str:
        # парсим запрос
        intent, entities = self.interpreter.parse(text)
        # предсказываем ответ бота
        return self.action_handler.process(intent, entities)



if __name__ == "__main__":
    # интерпретатор фраз т.е. превразает фразу в намерение и сущности
    interpreter = TwinInterpreter()
    # обработчик ответов бота
    action_handler = ActionHandler()
    # сам диалог
    dialog = DialogTracker(interpreter = interpreter, action_handler=action_handler)
    while True:
        with m as source:
            print("Говорите")
            audio = r.listen(source)
        query = r.recognize_google(audio, language="ru-RU")
        print("Вы сказали: " + query.lower())
        input_text = query
        output_text = dialog.predict(input_text)
        print("Бот:")
        print(output_text)
        engine.say(output_text)
        engine.runAndWait()
    pass